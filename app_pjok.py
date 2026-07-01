import streamlit as st
import pandas as pd
import openpyxl
import os

st.set_page_config(page_title="Asisten PJOK Digital", layout="centered")

st.title("🏃‍♂️ Asisten PJOK Digital: Presensi & Nilai")
st.caption("Aplikasi Web Terintegrasi Langsung dengan Spreadsheet Excel")

# Nama file Excel yang bertindak sebagai database
FILE_ABSEN = "Daftar_Hadir_PJOK_SD_Semester_1.xlsx"
FILE_NILAI = "Daftar_Nilai_PJOK_SD_Semester_1.xlsx"

# Fungsi untuk membaca data dari Excel
def load_data(file_path, sheet_name):
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=4)
        df = df[df['Nama Siswa'].notna()]
        df = df[df['Nama Siswa'] != 'Rata-Rata Kelas']
        return df
    else:
        st.error(f"Berkas '{file_path}' tidak ditemukan! Pastikan file Excel berada di folder yang sama dengan kode ini.")
        return None

# Fungsi untuk menulis data absensi ke Excel
def save_absensi_to_excel(file_path, sheet_name, kolom_pertemuan, status_absen):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    target_col = None
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=5, column=col).value
        if cell_val == kolom_pertemuan:
            target_col = col
            break
            
    if not target_col:
        st.error(f"Kolom {kolom_pertemuan} tidak ditemukan di file Excel.")
        return
        
    for row in range(7, ws.max_row + 1):
        nama_cell = ws.cell(row=row, column=3).value # Kolom C adalah Nama Siswa
        if nama_cell in status_absen:
            ws.cell(row=row, column=target_col, value=status_absen[nama_cell])
            
    wb.save(file_path)
    st.success(f"🎉 Presensi {kolom_pertemuan} berhasil disimpan langsung ke berkas Excel!")

# Navigasi Antarmuka HP
st.sidebar.header("⚙️ Navigasi")
pilihan_kelas = st.sidebar.selectbox("Pilih Kelas:", ["Kelas IV", "Kelas V", "Kelas VI"])
pilihan_fitur = st.sidebar.radio("Pilih Administrasi:", ["📊 Daftar Hadir (Presensi)", "💯 Daftar Nilai Formatif/Sumatif"])

if pilihan_fitur == "📊 Daftar Hadir (Presensi)":
    st.header(f"📅 Presensi Lapangan - {pilihan_kelas}")
    
    df_absen = load_data(FILE_ABSEN, pilihan_kelas)
    
    if df_absen is not None:
        list_pertemuan = [f"P-{i}" for i in range(1, 19)]
        pilihan_p = st.selectbox("Pilih Pertemuan:", list_pertemuan)
        
        st.info("💡 Pilih status kehadiran siswa, lalu klik Simpan di bawah.")
        
        with st.form("form_presensi_eksternal"):
            status_hari_ini = {}
            for idx, row_data in df_absen.iterrows():
                nama = row_data['Nama Siswa']
                nisn = row_data['NISN']
                
                col1, col2 = st.columns([2, 2])
                with col1:
                    st.write(f"**{nama}**")
                with col2:
                    val_sebelumnya = row_data.get(pilihan_p, "H")
                    if val_sebelumnya not in ["H", "S", "I", "A"]:
                        val_sebelumnya = "H"
                        
                    mapping = {"H": 0, "S": 1, "I": 2, "A": 3}
                    idx_default = mapping.get(val_sebelumnya, 0)
                    
                    status_hari_ini[nama] = st.radio(
                        f"Absen_{nama}", 
                        ["H", "S", "I", "A"], 
                        index=idx_default,
                        key=f"absen_{nama}", 
                        horizontal=True,
                        label_visibility="collapsed"
                    )
            
            submit = st.form_submit_button("💾 Simpan Presensi ke Excel")
            if submit:
                save_absensi_to_excel(FILE_ABSEN, pilihan_kelas, pilihan_p, status_hari_ini)

else:
    st.header(f"💯 Input Nilai Formatif & Sumatif - {pilihan_kelas}")
    st.info("💡 Seluruh nama siswa di atas ditarik otomatis dari file Excel Anda secara real-time!")
